import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from database import get_batch_records_for_export

def generate_excel_for_batch(batch_id: str) -> bytes:
    """
    Fetches all records for a batch from Supabase, structures them,
    and returns a beautifully styled Excel file as binary bytes.
    """
    # 1. Fetch raw joined records from DB
    raw_records = get_batch_records_for_export(batch_id)
    if not raw_records:
        raise ValueError(f"No records found in database for batch: {batch_id}")

    # 2. Flatten JSON structure for tabular presentation
    flat_records = []
    for rec in raw_records:
        data = rec.get("extracted_data") or {}
        
        # Supabase joins can sometimes return the joined model as a list
        if isinstance(data, list):
            data = data[0] if len(data) > 0 else {}
        elif data is None:
            data = {}

        flat_records.append({
            "Filename": rec.get("filename"),
            "Status": rec.get("status"),
            "Product Name": data.get("product_name") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Brand": data.get("brand") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Flavour": data.get("flavour") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Net Content": data.get("net_weight") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Category": data.get("product_type") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Barcode": data.get("barcode") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Ingredients": data.get("ingredients") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Expiry Date": data.get("expiry_date") or ("-" if rec.get("status") == "COMPLETED" else "N/A"),
            "Error Details": rec.get("error_message") or "-",
            "Processed At": rec.get("processed_at") or "-"
        })

    # 3. Create Pandas DataFrame
    df = pd.DataFrame(flat_records)

    # 4. Generate styled Excel in memory buffer
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extraction Report")
        
        # Get openpyxl objects to style the spreadsheet
        workbook = writer.book
        worksheet = writer.sheets["Extraction Report"]
        
        # Define modern styles
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald Green
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        row_font = Font(name="Segoe UI", size=10)
        error_font = Font(name="Segoe UI", size=10, color="DC2626") # Red text for errors
        
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )
        
        # Style Header Row
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        worksheet.row_dimensions[1].height = 28

        # Style Data Rows
        for row_idx in range(2, len(flat_records) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            status_cell = worksheet.cell(row=row_idx, column=2) # Column 2 is Status
            
            # Highlight Failed rows
            is_failed = status_cell.value == "FAILED"
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = error_font if (is_failed and col_idx in [2, 11]) else row_font
                cell.border = thin_border
                
                # Center align status and processed columns
                if col_idx in [2, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    # Clip columns that could contain very long text to prevent super-wide sheets
                    max_len = min(len(val), 40)
                    
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output.seek(0)
    return output.getvalue()
